from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Dict, List, Tuple, Optional

from flask import Flask, render_template, request, session, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
import os
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key")


# -----------------------------
# Нормативные пороги (MVP)
# Принцип: чем выше класс напряжения, тем строже нормы.
# Источник-идеология: IEC 60422 + практика эксплуатации (упрощенная модель для MVP).
# -----------------------------

@dataclass(frozen=True)
class ThresholdsMax:
    warn_max: float
    crit_max: float


@dataclass(frozen=True)
class ThresholdsMin:
    warn_min: float
    crit_min: float


# -----------------------------
# Профили
# -----------------------------
# Важно: для каждого профиля свой набор TH и (при желании) свои WEIGHTS.

TH_6_35: Dict[str, object] = {
    # Чем ниже, тем хуже (min)
    "bdv_kv": ThresholdsMin(warn_min=30.0, crit_min=25.0),
    "flash_c": ThresholdsMin(warn_min=140.0, crit_min=135.0),

    # Чем выше, тем хуже (max)
    "moisture_ppm": ThresholdsMax(warn_max=20.0, crit_max=30.0),
    "acid_mgkoh_g": ThresholdsMax(warn_max=0.10, crit_max=0.20),
    "tgdelta_pct": ThresholdsMax(warn_max=1.5, crit_max=2.5),
}

TH_110: Dict[str, object] = {
    "bdv_kv": ThresholdsMin(warn_min=40.0, crit_min=35.0),
    "flash_c": ThresholdsMin(warn_min=140.0, crit_min=135.0),

    "moisture_ppm": ThresholdsMax(warn_max=15.0, crit_max=25.0),
    "acid_mgkoh_g": ThresholdsMax(warn_max=0.08, crit_max=0.15),
    "tgdelta_pct": ThresholdsMax(warn_max=1.0, crit_max=2.0),
}

TH_220_330: Dict[str, object] = {
    "bdv_kv": ThresholdsMin(warn_min=50.0, crit_min=45.0),
    "flash_c": ThresholdsMin(warn_min=140.0, crit_min=135.0),

    "moisture_ppm": ThresholdsMax(warn_max=10.0, crit_max=20.0),
    "acid_mgkoh_g": ThresholdsMax(warn_max=0.05, crit_max=0.10),
    "tgdelta_pct": ThresholdsMax(warn_max=0.8, crit_max=1.5),
}

# Веса можно оставить общими, но теперь профили уже различаются порогами.
WEIGHTS_BASE: Dict[str, float] = {
    "moisture_ppm": 0.25,
    "bdv_kv": 0.25,
    "acid_mgkoh_g": 0.15,
    "tgdelta_pct": 0.15,
    "flash_c": 0.10,
    "impurities": 0.05,
    "water_extract": 0.05,
}

PROFILES = {
    "PWR_6_35": {
        "name": "Силовой трансформатор 6-35 кВ",
        "source": "IEC 60422 (упрощенная модель для MVP), отраслевые практики",
        "TH": TH_6_35,
        "WEIGHTS": WEIGHTS_BASE,
    },
    "PWR_110": {
        "name": "Силовой трансформатор 110 кВ",
        "source": "IEC 60422 (упрощенная модель для MVP), отраслевые практики",
        "TH": TH_110,
        "WEIGHTS": WEIGHTS_BASE,
    },
    "PWR_220_330": {
        "name": "Силовой трансформатор 220-330 кВ",
        "source": "IEC 60422 (упрощенная модель для MVP), отраслевые практики",
        "TH": TH_220_330,
        "WEIGHTS": WEIGHTS_BASE,
    },
}


# -----------------------------
# Оценка зон
# -----------------------------

def zone_max(value: float, t: ThresholdsMax) -> Tuple[str, float]:
    if value <= t.warn_max:
        return "Норма", 0.0
    if value <= t.crit_max:
        return "Предупреждение", 0.5
    return "Критично", 1.0


def zone_min(value: float, t: ThresholdsMin) -> Tuple[str, float]:
    if value >= t.warn_min:
        return "Норма", 0.0
    if value >= t.crit_min:
        return "Предупреждение", 0.5
    return "Критично", 1.0


def zone_impurities(value: str) -> Tuple[str, float]:
    v = (value or "").strip().lower()
    if v == "нет":
        return "Норма", 0.0
    if v in {"следы", "незначительно"}:
        return "Предупреждение", 0.5
    return "Критично", 1.0


def zone_water_extract(value: str) -> Tuple[str, float]:
    v = (value or "").strip().lower()
    if v == "нейтральная":
        return "Норма", 0.0
    if v in {"слабокислая", "слабокисл.", "слабокислая реакция"}:
        return "Предупреждение", 0.5
    return "Критично", 1.0


def compute_index(scores: Dict[str, float], weights: Dict[str, float]) -> float:
    total = 0.0
    wsum = 0.0
    for k, w in weights.items():
        if k in scores:
            total += scores[k] * w
            wsum += w
    if wsum == 0:
        return 0.0
    return round((total / wsum) * 100.0, 1)


def overall_status(index: float, any_critical: bool) -> str:
    # MVP правило: критично если есть хотя бы один критичный показатель или высокий индекс
    if any_critical or index >= 60:
        return "КРИТИЧЕСКОЕ"
    if index >= 25:
        return "ПОГРАНИЧНОЕ"
    return "НОРМА"


def build_recommendations(rows: List[dict]) -> List[str]:
    rec: List[str] = []
    by_name = {r["name"]: r for r in rows}

    moist = by_name["Влагосодержание (ppm)"]["zone"]
    bdv = by_name["Пробивное напряжение (кВ)"]["zone"]
    acid = by_name["Кислотное число (мг КОН/г)"]["zone"]
    tg = by_name["tg δ при 90°C (%)"]["zone"]
    imp = by_name["Механические примеси"]["zone"]
    wext = by_name["Реакция водной вытяжки"]["zone"]
    flash = by_name["Температура вспышки (°C)"]["zone"]

    if moist == "Критично" or bdv == "Критично":
        rec.append("Срочно выполнить сушку/дегазацию и фильтрацию масла (приоритет: высокий).")
        rec.append("Проверить герметичность трансформатора и состояние дыхательной системы/адсорбента.")
        rec.append("Повторить измерение пробивного напряжения после обработки масла.")

    if imp == "Критично":
        rec.append("Выполнить фильтрацию для удаления механических примесей; оценить возможный источник загрязнения.")

    if acid in {"Предупреждение", "Критично"} or tg in {"Предупреждение", "Критично"} or wext in {"Предупреждение", "Критично"}:
        rec.append("Рассмотреть регенерацию масла или частичную/полную замену при подтверждении устойчивого роста параметров старения.")
        rec.append("Рекомендуется контроль повторной пробы в ближайшие 2-4 недели (или по регламенту предприятия).")

    if flash in {"Предупреждение", "Критично"}:
        rec.append("Проверить пожаробезопасные характеристики масла; при снижении ниже нормы рассмотреть замену.")

    if not rec:
        rec.append("Параметры в пределах нормы. Продолжать эксплуатацию согласно регламенту, выполнять плановый контроль.")

    return rec


# -----------------------------
# Excel
# -----------------------------

def excel_from_result(result: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Заключение"

    bold = Font(bold=True)
    hfill = PatternFill("solid", fgColor="F4E9D3")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Заключение по пробе трансформаторного масла"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws["A3"] = "Нормативный профиль"
    ws["B3"] = result.get("profile_name") or "не указан"

    ws["A4"] = "Источник норм (справочно)"
    ws["B4"] = result.get("profile_source") or "не указан"

    ws["A5"] = "Трансформатор (опционально)"
    ws["B5"] = result.get("transformer_id") or "не указан"

    ws["A6"] = "Дата пробы"
    ws["B6"] = result.get("sample_date") or "не указана"

    ws["A7"] = "Интегральный индекс"
    ws["B7"] = f'{result.get("index_score", 0)} / 100'

    ws["A8"] = "Состояние"
    ws["B8"] = result.get("status") or ""

    for cell in ["A3", "A4", "A5", "A6", "A7", "A8"]:
        ws[cell].font = bold

    ws["A10"] = "Показатель"
    ws["B10"] = "Значение"
    ws["C10"] = "Оценка"
    ws["D10"] = "Пояснение"
    for c in ["A10", "B10", "C10", "D10"]:
        ws[c].font = bold
        ws[c].fill = hfill
        ws[c].border = border
        ws[c].alignment = Alignment(vertical="center")

    row_i = 11
    for r in result.get("rows", []):
        ws[f"A{row_i}"] = r["name"]
        ws[f"B{row_i}"] = f'{r["value"]} {r["unit"]}'.strip()
        ws[f"C{row_i}"] = r["zone"]
        ws[f"D{row_i}"] = r["comment"]
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row_i}"].border = border
            ws[f"{col}{row_i}"].alignment = Alignment(wrap_text=True, vertical="top")
        row_i += 1

    row_i += 1
    ws[f"A{row_i}"] = "Рекомендации"
    ws[f"A{row_i}"].font = bold
    ws.merge_cells(f"A{row_i}:D{row_i}")
    row_i += 1

    recs = result.get("recommendations", [])
    for i, rec in enumerate(recs, start=1):
        ws[f"A{row_i}"] = f"{i}."
        ws[f"B{row_i}"] = rec
        ws.merge_cells(f"B{row_i}:D{row_i}")
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row_i}"].alignment = Alignment(wrap_text=True, vertical="top")
        row_i += 1

    row_i += 1
    ws[f"A{row_i}"] = "Принадлежит Казанскому Государственному Энергетическому Университету. Создал Ахметов Айдар Русланович."
    ws.merge_cells(f"A{row_i}:D{row_i}")
    ws[f"A{row_i}"].alignment = Alignment(wrap_text=True)
    ws[f"A{row_i}"].font = Font(size=10)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 60

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# -----------------------------
# Helpers
# -----------------------------

def parse_float_field(field_name: str) -> float:
    raw = (request.form.get(field_name) or "").replace(",", ".").strip()
    if raw == "":
        raise ValueError(f"Поле '{field_name}' пустое.")
    try:
        return float(raw)
    except ValueError as e:
        raise ValueError(f"Поле '{field_name}' должно быть числом. Получено: '{raw}'") from e


# -----------------------------
# Routes
# -----------------------------

@app.get("/")
def index():
    profiles_for_ui = [{"id": k, "name": v["name"]} for k, v in PROFILES.items()]
    profiles_for_ui.sort(key=lambda x: x["name"])
    return render_template("index.html", profiles=profiles_for_ui)


@app.post("/evaluate")
def evaluate():
    # 1) Профиль (влияет на расчёт)
    profile_id = (request.form.get("profile_id") or "").strip()
    profile = PROFILES.get(profile_id) or PROFILES["PWR_110"]  # дефолт 110 кВ

    th = profile["TH"]
    weights = profile["WEIGHTS"]
    profile_name = profile["name"]
    profile_source = profile.get("source", "")

    # 2) Инфо поля (не влияют на расчёт)
    transformer_id = (request.form.get("transformer_id") or "").strip()
    sample_date = (request.form.get("sample_date") or "").strip()

    # 3) Показатели масла
    try:
        moisture_ppm = parse_float_field("moisture_ppm")
        bdv_kv = parse_float_field("bdv_kv")
        acid = parse_float_field("acid_mgkoh_g")
        tg = parse_float_field("tgdelta_pct")
        flash = parse_float_field("flash_c")
    except ValueError as err:
        return str(err), 400

    impurities = (request.form.get("impurities") or "нет").strip()
    water_extract = (request.form.get("water_extract") or "нейтральная").strip()

    rows: List[dict] = []
    scores: Dict[str, float] = {}
    any_critical = False

    z, s = zone_max(moisture_ppm, th["moisture_ppm"])
    rows.append({
        "name": "Влагосодержание (ppm)",
        "value": moisture_ppm,
        "unit": "ppm",
        "zone": z,
        "comment": "Рост влаги снижает электрическую прочность и ускоряет старение изоляции."
    })
    scores["moisture_ppm"] = s
    any_critical = any_critical or (z == "Критично")

    z, s = zone_min(bdv_kv, th["bdv_kv"])
    rows.append({
        "name": "Пробивное напряжение (кВ)",
        "value": bdv_kv,
        "unit": "кВ",
        "zone": z,
        "comment": "Низкое пробивное напряжение повышает риск пробоя изоляции."
    })
    scores["bdv_kv"] = s
    any_critical = any_critical or (z == "Критично")

    z, s = zone_max(acid, th["acid_mgkoh_g"])
    rows.append({
        "name": "Кислотное число (мг КОН/г)",
        "value": acid,
        "unit": "мг КОН/г",
        "zone": z,
        "comment": "Рост кислотного числа отражает старение масла и накопление продуктов окисления."
    })
    scores["acid_mgkoh_g"] = s
    any_critical = any_critical or (z == "Критично")

    z, s = zone_max(tg, th["tgdelta_pct"])
    rows.append({
        "name": "tg δ при 90°C (%)",
        "value": tg,
        "unit": "%",
        "zone": z,
        "comment": "Повышенный tg δ указывает на рост диэлектрических потерь и загрязнение/старение."
    })
    scores["tgdelta_pct"] = s
    any_critical = any_critical or (z == "Критично")

    z, s = zone_min(flash, th["flash_c"])
    rows.append({
        "name": "Температура вспышки (°C)",
        "value": flash,
        "unit": "°C",
        "zone": z,
        "comment": "Снижение температуры вспышки ухудшает пожаробезопасность."
    })
    scores["flash_c"] = s
    any_critical = any_critical or (z == "Критично")

    z, s = zone_impurities(impurities)
    rows.append({
        "name": "Механические примеси",
        "value": impurities,
        "unit": "",
        "zone": z,
        "comment": "Примеси снижают электрическую прочность и ускоряют старение масла."
    })
    scores["impurities"] = s
    any_critical = any_critical or (z == "Критично")

    z, s = zone_water_extract(water_extract)
    rows.append({
        "name": "Реакция водной вытяжки",
        "value": water_extract,
        "unit": "",
        "zone": z,
        "comment": "Кислая реакция может указывать на наличие кислотных продуктов старения."
    })
    scores["water_extract"] = s
    any_critical = any_critical or (z == "Критично")

    index_score = compute_index(scores, weights)
    status = overall_status(index_score, any_critical)
    recs = build_recommendations(rows)

    result = {
        "profile_id": profile_id or "PWR_110",
        "profile_name": profile_name,
        "profile_source": profile_source,
        "transformer_id": transformer_id,
        "sample_date": sample_date,
        "status": status,
        "index_score": index_score,
        "rows": rows,
        "recommendations": recs,
    }
    session["last_result"] = result

    return render_template(
        "result.html",
        profile_name=profile_name,
        transformer_id=transformer_id,
        sample_date=sample_date,
        status=status,
        index_score=index_score,
        rows=rows,
        recommendations=recs,
    )


@app.get("/export/xlsx")
def export_xlsx():
    result = session.get("last_result")
    if not result:
        return "Нет данных для экспорта. Сначала выполните расчет.", 400

    bio = excel_from_result(result)
    fname = "zaklyuchenie_transformatornoe_maslo.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
